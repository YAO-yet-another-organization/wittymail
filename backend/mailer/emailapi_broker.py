#!/usr/bin/env python
# coding=utf-8

import os, sys
import openpyxl
from wittymail_server import flask_app
import mailer.emailapi as emailapi
import re
import util.logger as logger
import tempfile
import pdb

# import ai.gender_guesser as gender_guesser

log = logger.get_logger(__name__)

global email_fodder_names
global email_fodder

email_fodder_names_STATUS_INDEX = -1 

extended_email_fodder_names = ['pronoun', 'email_subject', 'email_body', 'email_attachment', 'email_sent']

extended_email_fodder_names_PRONOUN_INDEX = -5
extended_email_fodder_names_EMAIL_SUBJECT_INDEX = -4
extended_email_fodder_names_EMAIL_BODY_INDEX = -3

global extended_email_fodder_names_EMAIL_ATTACHMENT_INDEX
extended_email_fodder_names_EMAIL_ATTACHMENT_INDEX = -2

global extended_email_fodder

# Index of the emailid in the email_fodder_names[]. The emailid will be used in the 'to' field
global EMAIL_FODDER_TO_INDEX
# Index of the emailid in the email_fodder_names[]. The emailid will be used in the 'cc' field
global EMAIL_FODDER_CC_INDEX

global email_from 

global attachment_dir
global attachment_index

# Extended fodder that will be appended to the fodder provided by the user (i.e. email_fodder[])
extended_default_email_fodder = ["his", None, None, None, False]

def _init():
    global email_fodder
    email_fodder = []  

    global email_fodder_names
    email_fodder_names = []

    global extended_email_fodder
    extended_email_fodder = []

    global EMAIL_FODDER_TO_INDEX
    EMAIL_FODDER_TO_INDEX = None

    global EMAIL_FODDER_CC_INDEX
    EMAIL_FODDER_CC_INDEX = None

    global attachment_index
    attachment_index = None

def save_fodder_to_file():
    wb=openpyxl.Workbook()
    ws_write = wb.active

    headers = []

    headers.extend(email_fodder_names)

    l = []
    l.append(headers)
    excel_data = l + email_fodder 

    for row in excel_data:
       ws_write.append(row)

    wb.save(filename="wittymail_excel.xlsx")
    return os.path.join(os.getcwd(), "wittymail_excel.xlsx")

# TODO Add check to detect extention type
def save_fodder_from_file(loc, email_fodder_names_template = None):
        global email_fodder
        global email_fodder_names
        email_fodder = []

        _init()
        # TODO Check if we should read wb_obj.active or wb_obj[0]. Do we have to close?
        wb_obj = openpyxl.load_workbook(loc, data_only=True)   
        sheet_obj = wb_obj.active

        # Get headers
        email_fodder_names = []
        for j in range(1, sheet_obj.max_column + 1):
            email_fodder_names.append(str(sheet_obj.cell(row = 1, column = j).value))
        
        for i in range(2, sheet_obj.max_row + 1):
            cells = []

            for j in range(1, len(email_fodder_names) + 1):
                cells.append(str(sheet_obj.cell(row = i, column = j).value))
        
            if all(c == 'None' for c in cells):
                break

            cells.append("E-mail pending") 
            email_fodder.append(cells)
            
        assert email_fodder_names_template == None or email_fodder_names == email_fodder_names_template, \
               "**** The headers in the excel sheet seems to be invalid ****"
        assert len(email_fodder) > 0,            \
               "**** There are no entries in the excel sheet! ****"

        email_fodder_names.append("Status")

def get_email_fodder_names():
        return email_fodder_names

def get_email_fodder():
        return email_fodder

def get_extended_email_fodder():
        return extended_email_fodder

# #{no} in _st will be replaced by l[no]
def template_to_str(st, l):
    log.debug('template_to_str() str = %s' % (st))

    str_to_replace = list(set(re.findall(r'#\d+', st)))

    def remove_hash(s):
        return s[1:]

    index = map(int, [remove_hash(s) for s in str_to_replace])
    index = list(index)

    for i in range(len(index)):
        st = st.replace(str_to_replace[i], l[index[i] - 1])

    log.debug('template_to_str() final str = %s' % st)
    return st

def save_extended_fodder(to_column, cc_column, subject_template, body_template):
    global EMAIL_FODDER_TO_INDEX
    global EMAIL_FODDER_CC_INDEX
    global extended_email_fodder

    EMAIL_FODDER_TO_INDEX = email_fodder_names.index(to_column)
    EMAIL_FODDER_CC_INDEX = email_fodder_names.index(cc_column)

    for r in email_fodder:
        gender = "his"

        extended_fodder_entry = []

        pronoun = "her" if gender == "female" else "his" 
        extended_fodder_entry.append(pronoun)

        extended_fodder_entry.append(template_to_str(subject_template, r))
        extended_fodder_entry.append(template_to_str(body_template, r))

        _attachments = r[attachment_index].split(',')
        attachments = [a + ".pdf" for a in _attachments]
        extended_fodder_entry.append(attachments)

        for a in attachments:
            if not os.path.isfile(os.path.join(flask_app.config['ATTACHMENTS_DIR'], a)):
                r[email_fodder_names_STATUS_INDEX] = "pdf not found"

        extended_fodder_entry.append(False)

        extended_email_fodder.append(extended_fodder_entry)

def save_attachment_dir(dir_loc):
        global attachment_dir 
        attachment_dir = dir_loc

def save_attachment_column(attachment_column):
    global attachment_index
    # pdb.set_trace()
    attachment_index = email_fodder_names.index(attachment_column)

def test_email(tos):
    ex = extended_email_fodder[0]
    log.info('Test mail to be sent to+cc %s' % (''.join(tos)))
    ccs = tos

    email_subject = ex[extended_email_fodder_names_EMAIL_SUBJECT_INDEX]
    email_body = ex[extended_email_fodder_names_EMAIL_BODY_INDEX]

    attachment_list = ex[extended_email_fodder_names_EMAIL_ATTACHMENT_INDEX]    
    attachments = [os.path.join(attachment_dir, a) for a in attachment_list]
    return emailapi.send_email(email_from, tos, email_subject, email_body, ccs, attachments)

def send_email(email_from, tos, subject, body, ccs, attachments):
    attachments = [os.path.join(attachment_dir, a) for a in attachments]
    return emailapi.send_email(email_from, tos, subject, body, ccs, attachments)

def set_login_details(username, password, server="smtp.gmail.com", port=587):
    assert username is not None and password is not None,\
    log.error('server, port number, username or password is None')

    global email_from

    email_from = username
    return emailapi.set_login_details(server, port, username, password)